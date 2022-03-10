from openpyxl import load_workbook
from users.models import Order, ProductInOrder
from os.path import join, dirname, realpath


folder = dirname(realpath(__file__))
input_file = join(folder, "orders.xlsx")


def table(order_id, name):
    wb = load_workbook(input_file)
    ws = wb['Лист1']

    order = Order.objects.get(id=order_id)
    product_order = ProductInOrder.objects.filter(order=order_id)

    ws['B4'] = "{lastname}".format(lastname=order.lastname)
    ws['B5'] = "{firstname}".format(firstname=order.firstname)
    ws['B6'] = "{city}".format(city=order.city)
    ws['B7'] = "{phone}".format(phone=order.phone)
    ws['B8'] = "{email}".format(email=order.email)

    ws['A10'] = "{order}".format(order=name)

    n = 13
    all_summ = 0

    for item in product_order:
        A = f"A{n}"
        B = f"B{n}"
        C = f"C{n}"
        D = f"D{n}"

        ws[A] = "{title}".format(title=item.product.title)
        ws[B] = "{nmb}".format(nmb=item.nmb)
        ws[C] = "{price}".format(price=item.price_per_item)
        ws[D] = "{summ}".format(summ=item.total_price)
        
        n += 1
        all_summ += item.total_price

    S = f"D{n+1}"
    ws[S] = "{all_s}".format(all_s=all_summ)

    try:
        output_file = f"{name}.xlsx"
        output_folder = join(folder, "orders")
        wb.save(join(output_folder, output_file))
        print('Document [Save]')
        
    except:
        print('Error [Save]')
